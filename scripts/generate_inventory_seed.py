from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
import unicodedata

ROOT = Path(__file__).resolve().parents[1]
INPUT_FILES = {
    "Monterrico": ROOT / "Monterrico.xlsx",
    "San Miguel": ROOT / "San Miguel.xlsx",
}
OUTPUT_SQL = ROOT / "supabase" / "INVENTORY_EXCEL_SEED.sql"


def sql_escape(value: str) -> str:
    return value.replace("'", "''")


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value or "item"


def normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return text.lower().strip()


def infer_category(name: str, comments: str | None) -> str:
    haystack = f"{name} {comments or ''}"
    normalized = normalize_text(haystack)

    rules = [
        ("VR", ["vr", "oculus", "vive", "rift", "quest", "realidad virtual", "headset"]),
        ("Cables", ["cable", "displayport", "hdmi", "adapter", "adaptador"]),
        ("Tablets", ["ipad", "tablet", "tab s", "galaxy tab", "tab "]),
        ("Celulares", ["iphone", "galaxi", "galaxy", "celular", "movil", "s20"]),
        ("Audio", ["parlante", "jbl", "soundstick", "life chat", "headset", "audifono"]),
        ("Cámaras", ["webcam", "camera", "camara", "kinect"]),
        ("Consolas", ["playstation", "xbox", "wii", "nintendo", "gaming", "racing wheel"]),
        ("Computadoras", ["laptop", "imac", "cpu", "dell", "core i", "pc"]),
        ("Monitores/TV", ["monitor", "televisor", "tv", "pantalla"]),
        ("Proyectores", ["proyector", "cine beam"]),
        ("Redes/IoT", ["router", "beacon", "raspberry", "arduino", "iot", "deep learning"]),
        ("Almacenamiento", ["disco", "hdd", "duro", "seaget", "seagate"]),
        ("Periféricos", ["control", "mouse", "teclado", "intous"]),
    ]

    for category, keywords in rules:
        if any(keyword in normalized for keyword in keywords):
            return category

    return "Otros"


def normalize_code(raw: str | None) -> str | None:
    if raw is None:
        return None
    code = str(raw).strip()
    if not code:
        return None
    lowered = code.lower()
    if lowered in {"sin codigo", "sin código", "s/c", "na", "n/a", "none"}:
        return None
    code = code.replace("'", "").strip()
    return code or None


def parse_sheet(ws) -> Tuple[List[dict], List[dict]]:
    products: Dict[Tuple[str, str], dict] = {}
    units: List[dict] = []
    product_code_counters: Dict[Tuple[str, str], Dict[str, int]] = {}

    source_lab = ws.title.strip()

    for r in range(1, ws.max_row + 1):
        qty = ws.cell(r, 2).value
        name = ws.cell(r, 3).value

        if not isinstance(qty, (int, float)):
            continue
        qty_int = int(qty)
        if qty_int <= 0:
            continue
        if not name or not str(name).strip():
            continue

        name_str = str(name).strip()
        comments = ws.cell(r, 4).value
        comments_str = str(comments).strip() if comments and str(comments).strip() else ""
        asset_code_raw = ws.cell(r, 5).value
        observation = ws.cell(r, 8).value if ws.max_column >= 8 else None
        if ws.max_column >= 9 and not observation:
            observation = ws.cell(r, 9).value

        description_parts = []
        if comments and str(comments).strip():
            description_parts.append(str(comments).strip())
        if observation and str(observation).strip():
            description_parts.append(f"Obs: {str(observation).strip()}")
        description = " | ".join(description_parts) if description_parts else ""

        semantic_category = infer_category(name_str, comments_str)
        key = (name_str.lower(), semantic_category.lower())
        if key not in products:
            products[key] = {
                "name": name_str,
                "category": semantic_category,
                "description": (f"Lab: {source_lab} | " + description) if description else f"Lab: {source_lab}",
            }

        asset_code = normalize_code(asset_code_raw)
        base_code = asset_code or f"AUTO-{slugify(source_lab)}-{slugify(name_str)}-{r}"
        counters = product_code_counters.setdefault(key, {})

        for i in range(qty_int):
            candidate_code = base_code if qty_int == 1 else f"{base_code}-{i + 1:02d}"
            seen_count = counters.get(candidate_code, 0) + 1
            counters[candidate_code] = seen_count
            unit_code = candidate_code if seen_count == 1 else f"{candidate_code}-REP{seen_count:02d}"

            units.append(
                {
                    "product_key": key,
                    "unit_code": unit_code,
                    "asset_code": asset_code,
                    "note": str(observation).strip() if observation and str(observation).strip() else None,
                }
            )

    return list(products.values()), units


def main() -> None:
    all_products: Dict[Tuple[str, str], dict] = {}
    all_units: List[dict] = []

    for campus, file_path in INPUT_FILES.items():
        if not file_path.exists():
            print(f"Aviso: no se encontró {file_path.name}. Se omite {campus}.")
            continue

        wb = openpyxl.load_workbook(file_path, data_only=True)

        for ws in wb.worksheets:
            products, units = parse_sheet(ws)
            for p in products:
                key = (p["name"].lower(), p["category"].lower())
                all_products.setdefault(key, p)

            for unit in units:
                unit["campus"] = campus
            all_units.extend(units)

    if not all_products and not all_units:
        raise RuntimeError(
            "No se encontraron datos en Monterrico.xlsx ni San Miguel.xlsx."
        )

    lines: List[str] = []
    lines.append("BEGIN;")
    lines.append("\n-- Seed generado desde Monterrico.xlsx y San Miguel.xlsx")

    lines.append(
        """
CREATE TEMP TABLE tmp_inventory_products (
  name text,
  category text,
  description text
);
""".strip()
    )

    if all_products:
        values = []
        for p in all_products.values():
            values.append(
                "('" + sql_escape(p["name"]) + "','" + sql_escape(p["category"]) + "','" + sql_escape(p["description"]) + "')"
            )
        lines.append("INSERT INTO tmp_inventory_products(name, category, description) VALUES\n" + ",\n".join(values) + ";")

    lines.append(
        """
INSERT INTO products(name, price, category, description, main_image, additional_images, featured, in_stock, stock)
SELECT
  t.name,
  0,
  t.category,
  NULLIF(t.description, ''),
  'https://placehold.co/600x400?text=UPC+Inventario',
  '{}',
  false,
  true,
  0
FROM tmp_inventory_products t
ON CONFLICT DO NOTHING;
""".strip()
    )

    lines.append(
        """
CREATE TEMP TABLE tmp_inventory_units (
  product_name text,
  product_category text,
  unit_code text,
    campus text,
  asset_code text,
  note text
);
""".strip()
    )

    if all_units:
        values = []
        for u in all_units:
            product_name = u["product_key"][0]
            product_category = u["product_key"][1]
            values.append(
                "('"
                + sql_escape(product_name)
                + "','"
                + sql_escape(product_category)
                + "','"
                + sql_escape(u["unit_code"])
                + "',"
                + "'"
                + sql_escape(u.get("campus") or "Monterrico")
                + "',"
                + ("NULL" if not u["asset_code"] else "'" + sql_escape(u["asset_code"]) + "'")
                + ","
                + ("NULL" if not u["note"] else "'" + sql_escape(u["note"]) + "'")
                + ")"
            )
        lines.append(
            "INSERT INTO tmp_inventory_units(product_name, product_category, unit_code, campus, asset_code, note) VALUES\n"
            + ",\n".join(values)
            + ";"
        )

    lines.append(
        """
INSERT INTO inventory_units(product_id, unit_code, campus, asset_code, current_note)
SELECT p.id, tu.unit_code, tu.campus, tu.asset_code, tu.note
FROM tmp_inventory_units tu
JOIN products p
  ON lower(p.name) = tu.product_name
 AND lower(p.category) = tu.product_category
ON CONFLICT (product_id, unit_code) DO UPDATE
SET campus = EXCLUDED.campus,
    asset_code = EXCLUDED.asset_code,
    current_note = EXCLUDED.current_note,
    updated_at = timezone('utc'::text, now());
""".strip()
    )

    lines.append(
        """
INSERT INTO inventory_unit_notes(unit_id, note)
SELECT iu.id, tu.note
FROM tmp_inventory_units tu
JOIN products p
  ON lower(p.name) = tu.product_name
 AND lower(p.category) = tu.product_category
JOIN inventory_units iu
  ON iu.product_id = p.id
 AND iu.unit_code = tu.unit_code
WHERE tu.note IS NOT NULL
ON CONFLICT DO NOTHING;
""".strip()
    )

    lines.append(
        """
UPDATE products p
SET stock = sub.cnt,
    in_stock = (sub.cnt > 0)
FROM (
  SELECT product_id, COUNT(*)::int AS cnt
  FROM inventory_units
  WHERE status = 'active'
  GROUP BY product_id
) sub
WHERE p.id = sub.product_id;
""".strip()
    )

    lines.append("COMMIT;")

    OUTPUT_SQL.write_text("\n\n".join(lines), encoding="utf-8")
    print(f"Seed SQL generado en: {OUTPUT_SQL}")
    print(f"Productos detectados: {len(all_products)}")
    print(f"Unidades detectadas: {len(all_units)}")


if __name__ == "__main__":
    main()
